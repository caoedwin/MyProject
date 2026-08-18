"""任务管理子系统 - 视图"""
import logging
from datetime import date

from django.db import models as db_models
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from system.utils import ok, fail, page_result
from system.permissions import HasPermission
from app01.models import User
from TaskManagement.models import (
    TaskCategory, Task, TaskApproval, TaskProgressRecord, TaskPerformance,
)
from TaskManagement.serializers import (
    TaskCategorySerializer, TaskListSerializer, TaskDetailSerializer,
    TaskCreateUpdateSerializer, TaskApprovalSerializer,
    TaskProgressRecordSerializer, TaskPerformanceSerializer,
    UserBriefSerializer,
)

logger = logging.getLogger(__name__)

# 角色编码常量
ROLE_ADMIN = 'DQA_C38_TaskManagement_admin'
ROLE_USER = 'DQA_C38_TaskManagement_users'
ROLE_QM = 'DQA_QM'   # 主管/经理
ROLE_PL = 'DQA_PL'   # 主管/组长
ROLE_TE = 'DQA_TE'   # 员工

# 主管角色集合
SUPERVISOR_ROLES = {ROLE_QM, ROLE_PL}


def _is_supervisor(user):
    """判断用户是否是主管（QM/PL）"""
    if user.is_superuser:
        return True
    role_codes = set(user.user_roles.values_list('role__code', flat=True))
    return bool(role_codes & SUPERVISOR_ROLES)


def _is_task_admin(user):
    """判断用户是否是子系统管理员"""
    if user.is_superuser:
        return True
    role_codes = set(user.user_roles.values_list('role__code', flat=True))
    return ROLE_ADMIN in role_codes


def _can_edit_all(user):
    """管理员/子系统管理员可以查看编辑所有任务"""
    return user.is_superuser or _is_task_admin(user)


def _can_edit_task(user, task):
    """判断用户是否可以编辑某个任务"""
    if _can_edit_all(user):
        return True
    if _is_supervisor(user):
        return True
    # 普通员工只能编辑自己创建/负责/参与的任务
    if task.created_by_id == user.id:
        return True
    if task.owner_id == user.id:
        return True
    if task.participants.filter(id=user.id).exists():
        return True
    return False


def _annotate_can_edit(user, tasks):
    """为任务列表中的每条记录标注 can_edit 字段"""
    result = []
    is_admin = _can_edit_all(user)
    is_supervisor = _is_supervisor(user)
    for task in tasks:
        data = task if isinstance(task, dict) else TaskListSerializer(task).data
        if is_admin or is_supervisor:
            data['can_edit'] = True
        else:
            data['can_edit'] = (
                task.created_by_id == user.id or
                task.owner_id == user.id or
                task.participants.filter(id=user.id).exists()
            )
        result.append(data)
    return result


# ============================================================
# 任务种类管理
# ============================================================
class TaskCategoryViewSet(viewsets.ModelViewSet):
    """任务种类 CRUD"""
    queryset = TaskCategory.objects.all()
    serializer_class = TaskCategorySerializer
    permission_classes = [IsAuthenticated, HasPermission]
    permission_required = 'TaskManagement:category'

    def get_queryset(self):
        qs = TaskCategory.objects.all()
        keyword = self.request.query_params.get('keyword')
        if keyword:
            qs = qs.filter(Q(name__icontains=keyword) | Q(code__icontains=keyword))
        return qs.order_by('sort', 'id')

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        return Response(ok(TaskCategorySerializer(qs, many=True).data))

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(created_by=request.user)
        return Response(ok(serializer.data, msg='创建成功'))

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(ok(serializer.data, msg='更新成功'))

    @action(detail=False, methods=['get'])
    def choices(self, request):
        """返回所有启用的任务种类（供前端下拉选择）"""
        categories = TaskCategory.objects.filter(status=True).order_by('sort', 'id')
        return Response(ok(TaskCategorySerializer(categories, many=True).data))


# ============================================================
# 任务管理（核心）
# ============================================================
class TaskViewSet(viewsets.ModelViewSet):
    """任务 CRUD + 审批 + 进度"""
    queryset = Task.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ('list',):
            return TaskListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return TaskCreateUpdateSerializer
        return TaskDetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Task.objects.select_related(
            'category', 'owner', 'created_by'
        ).prefetch_related('participants')

        # 管理员/子系统管理员：查看所有
        if _can_edit_all(user):
            pass
        # 主管：查看所有员工的任务
        elif _is_supervisor(user):
            pass
        # 普通员工：只看自己创建/负责/参与的
        else:
            qs = qs.filter(
                Q(created_by=user) | Q(owner=user) | Q(participants=user)
            ).distinct()

        # 筛选条件
        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        category_id = self.request.query_params.get('category')
        if category_id:
            qs = qs.filter(category_id=category_id)
        priority = self.request.query_params.get('priority')
        if priority:
            qs = qs.filter(priority=priority)
        difficulty = self.request.query_params.get('difficulty')
        if difficulty:
            qs = qs.filter(difficulty=difficulty)
        owner_id = self.request.query_params.get('owner')
        if owner_id:
            qs = qs.filter(owner_id=owner_id)
        keyword = self.request.query_params.get('keyword')
        if keyword:
            qs = qs.filter(
                Q(title__icontains=keyword) | Q(description__icontains=keyword)
            )

        # 排序
        ordering = self.request.query_params.get('ordering', '-created_at')
        return qs.order_by(ordering)

    def list(self, request, *args, **kwargs):
        """任务列表（分页）"""
        qs = self.filter_queryset(self.get_queryset())
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 100))
        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        tasks = qs[start:end]
        data = _annotate_can_edit(request.user, tasks)
        return Response(page_result(data, total, page, page_size))

    def retrieve(self, request, *args, **kwargs):
        """任务详情"""
        instance = self.get_object()
        # 权限检查
        if not _can_edit_all(request.user) and not _is_supervisor(request.user):
            if not (instance.created_by_id == request.user.id or
                    instance.owner_id == request.user.id or
                    instance.participants.filter(id=request.user.id).exists()):
                return Response(fail('无权查看该任务', code=403))
        serializer = TaskDetailSerializer(instance)
        data = serializer.data
        data['can_edit'] = _can_edit_task(request.user, instance)
        return Response(ok(data))

    def create(self, request, *args, **kwargs):
        """创建任务"""
        serializer = TaskCreateUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        task = serializer.save(created_by=request.user)
        return Response(ok(TaskDetailSerializer(task).data, msg='任务创建成功'))

    def update(self, request, *args, **kwargs):
        """更新任务"""
        instance = self.get_object()
        if not _can_edit_task(request.user, instance):
            return Response(fail('无权编辑该任务', code=403))
        # 审核通过的任务，员工只能编辑进度和状态
        if instance.status == Task.Status.APPROVED and not _can_edit_all(request.user) and not _is_supervisor(request.user):
            allowed_fields = {'progress', 'status', 'actual_end_date'}
            data = {k: v for k, v in request.data.items() if k in allowed_fields}
        else:
            data = request.data
        serializer = TaskCreateUpdateSerializer(instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        task = serializer.save(updated_by=request.user)
        return Response(ok(TaskDetailSerializer(task).data, msg='任务更新成功'))

    def destroy(self, request, *args, **kwargs):
        """删除任务"""
        instance = self.get_object()
        if not _can_edit_task(request.user, instance):
            return Response(fail('无权删除该任务', code=403))
        instance.delete()
        return Response(ok(msg='删除成功'))

    # ---------- 审批 ----------
    @action(detail=True, methods=['post'])
    def submit_approval(self, request, pk=None):
        """提交审核（员工 -> 主管）"""
        task = self.get_object()
        if task.status not in (Task.Status.DRAFT, Task.Status.REJECTED):
            return Response(fail('当前状态不可提交审核', code=400))
        if not _can_edit_task(request.user, task):
            return Response(fail('无权操作', code=403))
        approver_id = request.data.get('approver_id')
        if not approver_id:
            return Response(fail('请选择审批人', code=400))
        # 创建审批记录
        TaskApproval.objects.create(
            task=task,
            approver_id=approver_id,
            status=TaskApproval.ApprovalStatus.PENDING
        )
        task.status = Task.Status.PENDING
        task.save(update_fields=['status', 'updated_at'])
        return Response(ok(msg='已提交审核'))

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """审批通过（主管）"""
        task = self.get_object()
        if task.status != Task.Status.PENDING:
            return Response(fail('当前状态不可审批', code=400))
        if not _is_supervisor(request.user) and not _can_edit_all(request.user):
            return Response(fail('无权审批', code=403))
        comment = request.data.get('comment', '')
        # 更新审批记录
        approval = task.approvals.filter(
            status=TaskApproval.ApprovalStatus.PENDING
        ).first()
        if approval:
            approval.status = TaskApproval.ApprovalStatus.APPROVED
            approval.comment = comment
            approval.save()
        task.status = Task.Status.APPROVED
        task.save(update_fields=['status', 'updated_at'])
        return Response(ok(msg='审批通过'))

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """驳回（主管）"""
        task = self.get_object()
        if task.status != Task.Status.PENDING:
            return Response(fail('当前状态不可驳回', code=400))
        if not _is_supervisor(request.user) and not _can_edit_all(request.user):
            return Response(fail('无权审批', code=403))
        comment = request.data.get('comment', '')
        approval = task.approvals.filter(
            status=TaskApproval.ApprovalStatus.PENDING
        ).first()
        if approval:
            approval.status = TaskApproval.ApprovalStatus.REJECTED
            approval.comment = comment
            approval.save()
        task.status = Task.Status.REJECTED
        task.save(update_fields=['status', 'updated_at'])
        return Response(ok(msg='已驳回'))

    # ---------- 进度 ----------
    @action(detail=True, methods=['post'])
    def update_progress(self, request, pk=None):
        """更新进度"""
        task = self.get_object()
        if task.status not in (Task.Status.APPROVED, Task.Status.IN_PROGRESS):
            return Response(fail('当前状态不可更新进度', code=400))
        if not _can_edit_task(request.user, task):
            return Response(fail('无权操作', code=403))
        progress = request.data.get('progress')
        description = request.data.get('description', '')
        if progress is None:
            return Response(fail('请填写进度', code=400))
        try:
            progress = int(progress)
            if not (0 <= progress <= 100):
                raise ValueError
        except (ValueError, TypeError):
            return Response(fail('进度必须在0-100之间', code=400))
        # 记录进度变更
        TaskProgressRecord.objects.create(
            task=task, user=request.user,
            progress=progress, description=description
        )
        # 更新任务进度
        old_status = task.status
        task.progress = progress
        if progress > 0 and task.status == Task.Status.APPROVED:
            task.status = Task.Status.IN_PROGRESS
        if progress >= 100:
            task.status = Task.Status.PENDING_VERIFY
            task.actual_end_date = date.today()
        task.save(update_fields=['progress', 'status', 'actual_end_date', 'updated_at'])
        return Response(ok(msg='进度更新成功'))

    # ---------- 验证 ----------
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """验证完成（主管验证任务）"""
        task = self.get_object()
        if task.status != Task.Status.PENDING_VERIFY:
            return Response(fail('当前状态不可验证', code=400))
        if not _is_supervisor(request.user) and not _can_edit_all(request.user):
            return Response(fail('无权验证', code=403))
        task.status = Task.Status.COMPLETED
        task.actual_end_date = date.today()
        task.save(update_fields=['status', 'actual_end_date', 'updated_at'])
        return Response(ok(msg='验证通过'))

    # ---------- 批量操作 ----------
    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        """批量删除"""
        ids = request.data.get('ids', [])
        if not ids:
            return Response(fail('请选择要删除的任务', code=400))
        user = request.user
        if _can_edit_all(user):
            Task.objects.filter(id__in=ids).delete()
        elif _is_supervisor(user):
            Task.objects.filter(id__in=ids).delete()
        else:
            Task.objects.filter(id__in=ids, created_by=user).delete()
        return Response(ok(msg=f'成功删除 {len(ids)} 条任务'))

    # ---------- 导入导出 ----------
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Excel 批量导入任务"""
        import openpyxl
        from io import BytesIO
        file = request.FILES.get('file')
        if not file:
            return Response(fail('请上传文件', code=400))
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            success_count = 0
            duplicate_count = 0
            duplicates = []
            errors = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_data = dict(zip(headers, row))
                    title = row_data.get('任务标题') or row_data.get('title')
                    if not title:
                        continue
                    # 检查重复
                    if Task.objects.filter(title=title).exists():
                        duplicate_count += 1
                        duplicates.append(title)
                        continue
                    # 查找 category
                    cat_name = row_data.get('任务种类') or row_data.get('category')
                    category = None
                    if cat_name:
                        category = TaskCategory.objects.filter(
                            Q(name=cat_name) | Q(code=cat_name)
                        ).first()
                    # 查找 owner
                    owner_name = row_data.get('负责人') or row_data.get('owner')
                    owner = None
                    if owner_name:
                        owner = User.objects.filter(username=owner_name).first()
                    task = Task.objects.create(
                        title=title,
                        description=row_data.get('任务描述', '') or row_data.get('description', ''),
                        category=category,
                        priority=int(row_data.get('优先级', 3)) if row_data.get('优先级') else 3,
                        difficulty=int(row_data.get('难度等级', 3)) if row_data.get('难度等级') else 3,
                        benefit=row_data.get('效益') or row_data.get('benefit'),
                        owner=owner,
                        start_date=row_data.get('开始时间') or row_data.get('start_date'),
                        expected_end_date=row_data.get('预计结束时间') or row_data.get('expected_end_date'),
                        created_by=request.user,
                    )
                    success_count += 1
                except Exception as e:
                    errors.append(f'第{row_idx}行: {str(e)}')
            msg = f'成功导入 {success_count} 条'
            if duplicate_count > 0:
                msg += f'，重复 {duplicate_count} 条'
            if errors:
                msg += f'，错误 {len(errors)} 条'
            return Response(ok({
                'success_count': success_count,
                'duplicate_count': duplicate_count,
                'duplicates': duplicates,
                'errors': errors,
            }, msg=msg))
        except Exception as e:
            return Response(fail(f'导入失败: {str(e)}', code=400))

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Excel 导出任务列表"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse
        qs = self.filter_queryset(self.get_queryset())
        wb = Workbook()
        ws = wb.active
        ws.title = '任务列表'
        headers = [
            '任务标题', '任务描述', '任务种类', '优先级', '难度等级',
            '效益/积分', '负责人', '参与人', '开始时间', '预计结束时间',
            '实际结束时间', '任务进度', '任务状态', '创建人', '创建时间', '更新时间',
        ]
        header_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row_idx, task in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=task.title)
            ws.cell(row=row_idx, column=2, value=task.description)
            ws.cell(row=row_idx, column=3, value=task.category.name if task.category else '')
            ws.cell(row=row_idx, column=4, value=task.get_priority_display())
            ws.cell(row=row_idx, column=5, value=task.get_difficulty_display())
            ws.cell(row=row_idx, column=6, value=float(task.benefit) if task.benefit else '')
            ws.cell(row=row_idx, column=7, value=task.owner.username if task.owner else '')
            ws.cell(row=row_idx, column=8, value=', '.join(task.participants.values_list('username', flat=True)))
            ws.cell(row=row_idx, column=9, value=task.start_date)
            ws.cell(row=row_idx, column=10, value=task.expected_end_date)
            ws.cell(row=row_idx, column=11, value=task.actual_end_date)
            ws.cell(row=row_idx, column=12, value=f'{task.progress}%')
            ws.cell(row=row_idx, column=13, value=task.get_status_display())
            ws.cell(row=row_idx, column=14, value=task.created_by.username if task.created_by else '')
            ws.cell(row=row_idx, column=15, value=task.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_idx, column=16, value=task.updated_at.strftime('%Y-%m-%d %H:%M'))
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=tasks.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def template_download(self, request):
        """下载导入模板"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        wb = Workbook()
        ws = wb.active
        ws.title = '任务导入模板'
        headers = [
            '任务标题', '任务描述', '任务种类', '优先级', '难度等级',
            '效益/积分', '负责人', '开始时间', '预计结束时间',
        ]
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        # 添加示例数据
        ws.cell(row=2, column=1, value='示例：周报整理')
        ws.cell(row=2, column=2, value='整理本周工作周报')
        ws.cell(row=2, column=3, value='日常任务')
        ws.cell(row=2, column=4, value='3（中）')
        ws.cell(row=2, column=5, value='3（中）')
        ws.cell(row=2, column=6, value='10')
        ws.cell(row=2, column=7, value='zhangsan')
        ws.cell(row=2, column=8, value='2026-08-18')
        ws.cell(row=2, column=9, value='2026-08-25')
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=task_template.xlsx'
        wb.save(response)
        return response


# ============================================================
# 审批记录
# ============================================================
class TaskApprovalViewSet(viewsets.ReadOnlyModelViewSet):
    """审批记录查询"""
    queryset = TaskApproval.objects.all()
    serializer_class = TaskApprovalSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = TaskApproval.objects.select_related('task', 'approver')
        if not _can_edit_all(user):
            qs = qs.filter(Q(approver=user) | Q(task__created_by=user))
        task_id = self.request.query_params.get('task')
        if task_id:
            qs = qs.filter(task_id=task_id)
        return qs.order_by('-created_at')

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        data = TaskApprovalSerializer(qs[start:end], many=True).data
        return Response(page_result(data, total, page, page_size))


# ============================================================
# 绩效积分
# ============================================================
class TaskPerformanceViewSet(viewsets.ModelViewSet):
    """绩效积分管理"""
    queryset = TaskPerformance.objects.all()
    serializer_class = TaskPerformanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = TaskPerformance.objects.select_related('task', 'evaluated_by')
        if not _can_edit_all(user) and not _is_supervisor(user):
            qs = qs.filter(task__owner=user)
        return qs.order_by('-created_at')

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        data = TaskPerformanceSerializer(qs[start:end], many=True).data
        return Response(page_result(data, total, page, page_size))

    def create(self, request, *args, **kwargs):
        """创建/更新绩效评分"""
        if not _is_supervisor(request.user) and not _can_edit_all(request.user):
            return Response(fail('无权评分', code=403))
        task_id = request.data.get('task')
        if not task_id:
            return Response(fail('请选择任务', code=400))
        # 检查任务是否已完成
        try:
            task = Task.objects.get(pk=task_id)
        except Task.DoesNotExist:
            return Response(fail('任务不存在', code=404))
        if task.status != Task.Status.COMPLETED:
            return Response(fail('只能对已完成的任务评分', code=400))
        # 创建或更新
        performance, created = TaskPerformance.objects.update_or_create(
            task_id=task_id,
            defaults={
                'base_score': request.data.get('base_score', 0),
                'priority_bonus': request.data.get('priority_bonus', 0),
                'time_bonus': request.data.get('time_bonus', 0),
                'quality_score': request.data.get('quality_score', 0),
                'total_score': request.data.get('total_score', 0),
                'comment': request.data.get('comment', ''),
                'evaluated_by': request.user,
            }
        )
        performance.calculate_total()
        performance.save()
        msg = '评分创建成功' if created else '评分更新成功'
        return Response(ok(TaskPerformanceSerializer(performance).data, msg=msg))


# ============================================================
# 统计
# ============================================================
class TaskStatsViewSet(viewsets.GenericViewSet):
    """任务统计"""
    permission_classes = [IsAuthenticated]

    def _get_base_queryset(self):
        user = self.request.user
        if _can_edit_all(user) or _is_supervisor(user):
            return Task.objects.all()
        return Task.objects.filter(
            Q(created_by=user) | Q(owner=user) | Q(participants=user)
        ).distinct()

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """任务总览统计"""
        qs = self._get_base_queryset()
        total = qs.count()
        # 按状态统计
        by_status = {}
        for status_val, status_label in Task.Status.choices:
            count = qs.filter(status=status_val).count()
            by_status[status_label] = count
        # 按优先级统计
        by_priority = {}
        for priority_val, priority_label in Task.Priority.choices:
            count = qs.filter(priority=priority_val).count()
            by_priority[priority_label] = count
        # 按种类统计
        by_category = list(
            qs.values('category__name').annotate(count=Count('id')).order_by('-count')
        )
        # 按难度统计
        by_difficulty = {}
        for diff_val, diff_label in Task.Difficulty.choices:
            count = qs.filter(difficulty=diff_val).count()
            by_difficulty[diff_label] = count
        return Response(ok({
            'total': total,
            'by_status': by_status,
            'by_priority': by_priority,
            'by_category': by_category,
            'by_difficulty': by_difficulty,
        }))

    @action(detail=False, methods=['get'])
    def by_person(self, request):
        """按人员统计任务"""
        qs = self._get_base_queryset()
        # 按负责人统计
        by_owner = list(
            qs.values('owner__username', 'owner__nickname').annotate(
                total=Count('id'),
                completed=Count('id', filter=Q(status=Task.Status.COMPLETED)),
                in_progress=Count('id', filter=Q(status=Task.Status.IN_PROGRESS)),
            ).order_by('-total')
        )
        return Response(ok({'by_owner': by_owner}))

    @action(detail=False, methods=['get'])
    def performance_summary(self, request):
        """绩效汇总统计"""
        if not _can_edit_all(request.user) and not _is_supervisor(request.user):
            return Response(fail('无权查看绩效统计', code=403))
        qs = TaskPerformance.objects.all()
        by_person = list(
            qs.values('task__owner__username', 'task__owner__nickname').annotate(
                sum_score=Sum('total_score'),
                avg_score=Avg('total_score'),
                task_count=Count('id'),
            ).order_by('-sum_score')
        )
        return Response(ok({'by_person': by_person}))

    @action(detail=False, methods=['get'])
    def users(self, request):
        """获取可选用户列表（供负责人/参与人选择）"""
        users = User.objects.filter(status=1)
        keyword = request.query_params.get('keyword')
        if keyword:
            users = users.filter(
                Q(username__icontains=keyword) | Q(nickname__icontains=keyword)
            )
        users = users[:50]
        return Response(ok(UserBriefSerializer(users, many=True).data))

    @action(detail=False, methods=['get'])
    def qm_users(self, request):
        """获取所有主管（DQA_QM 角色）用户列表"""
        from system.models import UserRole
        qm_user_ids = UserRole.objects.filter(
            role__code='DQA_QM'
        ).values_list('user_id', flat=True)
        users = User.objects.filter(id__in=qm_user_ids, status=1)
        return Response(ok(UserBriefSerializer(users, many=True).data))